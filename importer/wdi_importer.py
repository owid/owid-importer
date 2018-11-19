import sys
import os
import hashlib
import json
import logging
import requests
import unidecode
import shutil
import time
import zipfile
# allow imports from parent directory
sys.path.insert(1, os.path.join(sys.path[0], '..'))
import grapher_admin.wsgi
from openpyxl import load_workbook
from grapher_admin.models import Entity, Dataset, DatasetTag, Source, Variable, Tag, DataValue, ChartDimension # rewrite removed DatasetSubcategory # rewrite removed DatasetCategory # rewrite removed VariableType
from importer.models import ImportHistory, AdditionalCountryInfo
from country_name_tool.models import CountryName
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from django.urls import reverse
from grapher_admin.views import write_dataset_csv

from importer_utils import file_checksum, extract_short_unit, get_row_values, starts_with, default


source_description = {
    'dataPublishedBy': "World Bank – World Development Indicators",
    'link': "http://data.worldbank.org/data-catalog/world-development-indicators",
    'retrievedDate': timezone.now().strftime("%d-%B-%y")
}

DATASET_NAMESPACE = 'wdi'
WDI_TAG_NAME = 'World Development Indicators'  # set the name of the root category of all data that will be imported by this script
WDI_ZIP_FILE_URL = 'http://databank.worldbank.org/data/download/WDI_excel.zip'
WDI_DOWNLOADS_PATH = settings.BASE_DIR + '/data/wdi_downloads/'

# The column headers we expect the sheets to have.
# We will only check that the headers begin with the columns listed here, if
# there are additional columns, that's fine and it shouldn't affect our script.
SERIES_EXPECTED_HEADERS = ('Series Code', 'Topic', 'Indicator Name', 'Short definition', 'Long definition', 'Unit of measure', 'Periodicity', 'Base Period', 'Other notes', 'Aggregation method', 'Limitations and exceptions', 'Notes from original source', 'General comments', 'Source', 'Statistical concept and methodology', 'Development relevance', 'Related source links', 'Other web links', 'Related indicators', 'License Type')
DATA_EXPECTED_HEADERS = ('Country Name', 'Country Code', 'Indicator Name', 'Indicator Code', '1960')
COUNTRY_EXPECTED_HEADERS = ('Country Code', 'Short Name', 'Table Name', 'Long Name', '2-alpha code', 'Currency Unit', 'Special Notes', 'Region', 'Income Group', 'WB-2 code', 'National accounts base year', 'National accounts reference year', 'SNA price valuation', 'Lending category', 'Other groups', 'System of National Accounts', 'Alternative conversion factor', 'PPP survey year', 'Balance of Payments Manual in use', 'External debt Reporting status', 'System of trade', 'Government Accounting concept', 'IMF data dissemination standard', 'Latest population census', 'Latest household survey', 'Source of most recent Income and expenditure data', 'Vital registration complete', 'Latest agricultural census', 'Latest industrial data', 'Latest trade data')

logger = logging.getLogger('importer')
start_time = time.time()

def terminate(message):
    logger.error(message)
    sys.exit(1)

# Extract indicator from row in Series worksheet
def get_indicator_from_row(row):
    values = get_row_values(row)
    code = values[0].upper().strip()
    indicator = {
        'code': code,
        'category': values[1].split(':')[0],
        'name': values[2],
        'description': values[4],
        'unitofmeasure': default(values[5], ''),
        'short_unit': None, # we will derive it later
        'limitations': default(values[10], ''),
        'sourcesnotes': default(values[11], ''),
        'comments': default(values[12], ''),
        'source': values[13],
        'concept': default(values[14], ''),
        'sourcelinks': default(values[16], ''),
        'weblinks': default(values[17], ''),
        'saved': False
    }
    # if no unit is specified, try to derive it from the name
    if not indicator['unitofmeasure'] and '(' in indicator['name'] and ')' in indicator['name']:
        indicator['unitofmeasure'] = indicator['name'][
            indicator['name'].rfind('(') + 1:
            indicator['name'].rfind(')')
        ]
    # derive the short unit
    indicator['short_unit'] = extract_short_unit(indicator['unitofmeasure'])
    return indicator

# Create a directory for holding the downloads.
if not os.path.exists(WDI_DOWNLOADS_PATH):
    os.makedirs(WDI_DOWNLOADS_PATH)

# logger.info("Getting the zip file")
# request_header = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
# r = requests.get(WDI_ZIP_FILE_URL, stream=True, headers=request_header)
# if r.ok:
#     with open(WDI_DOWNLOADS_PATH + 'wdi.zip', 'wb') as out_file:
#         shutil.copyfileobj(r.raw, out_file)
#     logger.info("Saved the zip file to disk.")
#     z = zipfile.ZipFile(WDI_DOWNLOADS_PATH + 'wdi.zip')
#     excel_filepath = WDI_DOWNLOADS_PATH + z.namelist()[0]  # there should be only one file inside the zipfile, so we will load that one
#     z.extractall(WDI_DOWNLOADS_PATH)
#     r = None  # we do not need the request anymore
#     logger.info("Successfully extracted the zip file")
# else:
#     logger.error("The file could not be downloaded. Stopping the script...")
#     sys.exit("Could not download file.")

import_history = ImportHistory.objects.filter(import_type=DATASET_NAMESPACE)

excel_filepath = WDI_DOWNLOADS_PATH + "WDIEXCEL.xlsx"

with transaction.atomic():
    # if wdi imports were never performed
    if not import_history:
        terminate("""The WDI data import has not been run before.

        The part of the script that deals with a fresh import is outdated and needs some fixes applied.
        """)

        logger.info("This is the very first WDI data import.")

        wb = load_workbook(excel_filepath, read_only=True)

        series_ws = wb['Series']
        data_ws = wb['Data']
        country_ws = wb['Country']

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0   # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        global_cat = {}  # global catalog of indicators

        # data in the worksheets is not loaded into memory at once, that causes RAM to quickly fill up
        # instead, we go through each row and cell one-by-one, looking at each piece of data separately
        # this has the disadvantage of needing to traverse the worksheet several times, if we need to look up some rows/cells again

        for row in series_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        global_cat[cell.value.upper().strip()] = {}
                        indicatordict = global_cat[cell.value.upper().strip()]
                    if column_number == 2:
                        indicatordict['category'] = cell.value.split(':')[0]
                    if column_number == 3:
                        indicatordict['name'] = cell.value
                    if column_number == 5:
                        indicatordict['description'] = cell.value
                    if column_number == 6:
                        if cell.value:
                            indicatordict['unitofmeasure'] = cell.value
                        else:
                            if '(' not in indicatordict['name']:
                                indicatordict['unitofmeasure'] = ''
                            else:
                                indicatordict['unitofmeasure'] = indicatordict['name'][
                                                                 indicatordict['name'].rfind('(') + 1:indicatordict[
                                                                     'name'].rfind(')')]
                    if column_number == 11:
                        if cell.value:
                            indicatordict['limitations'] = cell.value
                        else:
                            indicatordict['limitations'] = ''
                    if column_number == 12:
                        if cell.value:
                            indicatordict['sourcenotes'] = cell.value
                        else:
                            indicatordict['sourcenotes'] = ''
                    if column_number == 13:
                        if cell.value:
                            indicatordict['comments'] = cell.value
                        else:
                            indicatordict['comments'] = ''
                    if column_number == 14:
                        indicatordict['source'] = cell.value
                    if column_number == 15:
                        if cell.value:
                            indicatordict['concept'] = cell.value
                        else:
                            indicatordict['concept'] = ''
                    if column_number == 17:
                        if cell.value:
                            indicatordict['sourcelinks'] = cell.value
                        else:
                            indicatordict['sourcelinks'] = ''
                    if column_number == 18:
                        if cell.value:
                            indicatordict['weblinks'] = cell.value
                        else:
                            indicatordict['weblinks'] = ''
                    indicatordict['saved'] = False

            column_number = 0

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = Tag.objects.values('name') # rewrite removed DatasetCategory
        existing_categories_list = {item['name'] for item in existing_categories}

        if WDI_TAG_NAME not in existing_categories_list:
            parent_tag = Tag(name=WDI_TAG_NAME, is_bulk_import=True) # rewrite removed DatasetCategory
            parent_tag.save()
            logger.info("Inserting a category %s." % WDI_TAG_NAME.encode('utf8'))

        else:
            parent_tag = Tag.objects.get(name=WDI_TAG_NAME) # rewrite removed DatasetCategory

        existing_subcategories = Tag.objects.filter(parent_id=parent_tag).values('name') # rewrite removed DatasetSubcategory
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        wdi_categories_list = []

        for key, value in category_vars.items():
            wdi_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = Tag(name=key, parent_id=parent_tag) # rewrite removed DatasetSubcategory
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'] for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        row_number = 0
        for row in country_ws.rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        country_code = cell.value
                    if column_number == 3:
                        country_name = cell.value
                    if column_number == 7:
                        country_special_notes = cell.value
                    if column_number == 8:
                        country_region = cell.value
                    if column_number == 9:
                        country_income_group = cell.value
                    if column_number == 24:
                        country_latest_census = cell.value
                    if column_number == 25:
                        country_latest_survey = cell.value
                    if column_number == 26:
                        country_recent_income_source = cell.value
                    if column_number == 30:
                        entity_info = AdditionalCountryInfo()
                        entity_info.country_code = country_code
                        entity_info.country_name = country_name
                        entity_info.country_wb_region = country_region
                        entity_info.country_wb_income_group = country_income_group
                        entity_info.country_special_notes = country_special_notes
                        entity_info.country_latest_census = country_latest_census
                        entity_info.country_latest_survey = country_latest_survey
                        entity_info.country_recent_income_source = country_recent_income_source
                        entity_info.save()
                        if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                            newentity = Entity.objects.get(name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                        elif country_name in existing_entities_list:
                            newentity = Entity.objects.get(name=country_name)
                        else:
                            newentity = Entity(name=country_name, validated=False)
                            newentity.save()
                            logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                        country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, entityId, variableId) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []
        datasets_list = []
        for category in wdi_categories_list:
            newdataset = Dataset(name='World Development Indicators - ' + category,
                                 description='This is a dataset imported by the automated fetcher',
                                 namespace=DATASET_NAMESPACE) # rewrite removed DatasetSubcategory
            newdataset.save()
            dataset_tag = DatasetTag(dataset_id=newdataset, tag_id=parent_tag)
            dataset_tag.save()
            datasets_list.append(newdataset)
            logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            row_number = 0
            for row in data_ws.rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 1:
                            country_name = cell.value
                        if column_number == 2:
                            country_code = cell.value
                        if column_number == 3:
                            indicator_name = cell.value
                        if column_number == 4:
                            indicator_code = cell.value.upper().strip()
                        if column_number > 4 and column_number <= last_available_year - 1960 + 5:
                            if cell.value or cell.value == 0:
                                data_values.append({'value': cell.value, 'year': 1960 - 5 + column_number})
                        if column_number > 4 and column_number == last_available_year - 1960 + 5:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    if not global_cat[indicator_code]['saved']:
                                        source_description['additionalInfo'] = "Definitions and characteristics of countries and other territories: " + "https://ourworldindata.org" + reverse("servewdicountryinfo") + "\n"
                                        source_description['additionalInfo'] += "Limitations and exceptions:\n" + global_cat[indicator_code]['limitations'] + "\n" if global_cat[indicator_code]['limitations'] else ''
                                        source_description['additionalInfo'] += "Notes from original source:\n" + global_cat[indicator_code]['sourcenotes'] + "\n" if global_cat[indicator_code]['sourcenotes'] else ''
                                        source_description['additionalInfo'] += "General comments:\n" + global_cat[indicator_code]['comments'] + "\n" if global_cat[indicator_code]['comments'] else ''
                                        source_description['additionalInfo'] += "Statistical concept and methodology:\n" + global_cat[indicator_code]['concept'] if global_cat[indicator_code]['concept'] else ''
                                        source_description['additionalInfo'] += "Related source links:\n" + global_cat[indicator_code]['sourcelinks'] + "\n" if global_cat[indicator_code]['sourcelinks'] else ''
                                        source_description['additionalInfo'] += "Other web links:\n" + global_cat[indicator_code]['weblinks'] + "\n" if global_cat[indicator_code]['weblinks'] else ''
                                        source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                        if 'iea.org' in json.dumps(source_description).lower() or 'iea stat' in json.dumps(source_description).lower() or 'iea 2014' in json.dumps(source_description).lower():
                                            source_description['dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                        else:
                                            source_description['dataPublishedBy'] = 'World Bank – World Development Indicators'
                                        newsource = Source(name='World Bank – WDI: ' + global_cat[indicator_code]['name'],
                                                           description=json.dumps(source_description),
                                                           datasetId=newdataset)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))
                                        s_unit = extract_short_unit(global_cat[indicator_code]['unitofmeasure'])
                                        newvariable = Variable(name=global_cat[indicator_code]['name'], unit=global_cat[indicator_code]['unitofmeasure'] if global_cat[indicator_code]['unitofmeasure'] else '', short_unit=s_unit, description=global_cat[indicator_code]['description'],
                                                               code=indicator_code, timespan='1960-' + str(last_available_year), datasetId=newdataset, sourceId=newsource) # rewrite removed VariableType
                                        newvariable.save()
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        global_cat[indicator_code]['saved'] = True
                                    else:
                                        newvariable = global_cat[indicator_code]['variable_object']
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'], country_name_entity_ref[country_code].pk, newvariable.pk))
                                    if len(data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []

                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        newimport = ImportHistory(import_type=DATASET_NAMESPACE, import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Initial import of WDI',
                                  import_state=json.dumps({'file_hash': file_checksum(WDI_DOWNLOADS_PATH + 'wdi.zip')}))
        newimport.save()
        # for dataset in datasets_list:
            # write_dataset_csv(dataset.pk, dataset.name, None, 'wdi_fetcher', '')
        logger.info("Import complete.")

    else:
        last_import = import_history.last()

        if json.loads(last_import.import_state)['file_hash'] == file_checksum(WDI_DOWNLOADS_PATH + 'wdi.zip'):
            logger.info('No updates available.')
            sys.exit(0)

        logger.info('New data is available.')

        # ======================================================================
        # Load the worksheets we will need and check that they have the columns
        # we expect.
        # ======================================================================

        wb = load_workbook(excel_filepath, read_only=True)

        series_ws_rows = wb['Series'].rows
        data_ws_rows = wb['Data'].rows
        country_ws_rows = wb['Country'].rows

        series_headers = get_row_values(next(series_ws_rows))
        data_headers = get_row_values(next(data_ws_rows))
        country_headers = get_row_values(next(country_ws_rows))

        if not starts_with(series_headers, SERIES_EXPECTED_HEADERS):
            terminate("Headers mismatch on 'Series' worksheet")
        if not starts_with(data_headers, DATA_EXPECTED_HEADERS):
            terminate("Headers mismatch on 'Data' worksheet")
        if not starts_with(country_headers, COUNTRY_EXPECTED_HEADERS):
            terminate("Headers mismatch on 'Country' worksheet")

        # ======================================================================
        # Initialise the data structures to track the state of the import.
        # ======================================================================

        deleted_indicators = {}  # This is used to keep track which variables' data values were already deleted before writing new values
        global_cat = {}  # global catalog of indicators

        available_variables = Variable.objects.filter(datasetId__in=Dataset.objects.filter(namespace=DATASET_NAMESPACE))
        available_variables_codes = [var['code'] for var in available_variables.values('code')]

        chart_dimension_var_ids = {
            item['variableId']
            for item in ChartDimension.objects.all().values('variableId').distinct()
        }

        existing_variables_ids = [item['id'] for item in available_variables.values('id')]
        existing_variables_id_code = {item['id']: item['code'] for item in available_variables.values('id', 'code')}
        existing_variables_code_id = {item['code']: item['id'] for item in available_variables.values('id', 'code')}

        # we will not be deleting any variables that are currently being used by charts
        var_codes_being_used = [
            existing_variables_id_code[var_id]
            for var_id in existing_variables_ids
            if var_id in chart_dimension_var_ids
        ]

        # ======================================================================
        # Initialise the data structures to track the state of the import.
        # ======================================================================

        column_number = 0  # this will be reset to 0 on each new row
        row_number = 0  # this will be reset to 0 if we switch to another worksheet, or start reading the worksheet from the beginning one more time

        # Data in the worksheets is read row by row, to avoid going exceeding
        # the limits of the heap and crashing the program.

        for row in series_ws_rows:
            indicator = get_indicator_from_row(row)
            # store the indicator in the global catalog
            global_cat[indicator['code']] = indicator

        new_var_codes = set(global_cat.keys())

        var_codes_to_add = list(new_var_codes.difference(available_variables_codes))
        newly_added_var_codes = list(new_var_codes.difference(available_variables_codes))
        var_codes_to_delete = list(set(available_variables_codes).difference(new_var_codes).difference(var_codes_being_used))

        for var_code in var_codes_to_delete:
            logger.info("Deleting data values for the variable: %s" % var_code.encode('utf8'))
            while DataValue.objects.filter(variableId__pk=existing_variables_code_id[var_code]).first():
                with connection.cursor() as c:  # if we don't limit the deleted values, the db might just hang
                    c.execute('DELETE FROM %s WHERE variableId = %s LIMIT 10000;' %
                                (DataValue._meta.db_table, existing_variables_code_id[var_code]))
            source_object = Variable.objects.get(code=var_code, datasetId__in=Dataset.objects.filter(namespace=DATASET_NAMESPACE)).sourceId
            Variable.objects.get(code=var_code, datasetId__in=Dataset.objects.filter(namespace=DATASET_NAMESPACE)).delete()
            logger.info("Deleting the variable: %s" % var_code.encode('utf8'))
            logger.info("Deleting the source: %s" % source_object.name.encode('utf8'))
            source_object.delete()

        category_vars = {}  # categories and their corresponding variables

        for key, value in global_cat.items():
            if value['category'] in category_vars:
                category_vars[value['category']].append(key)
            else:
                category_vars[value['category']] = []
                category_vars[value['category']].append(key)

        existing_categories = Tag.objects.values('name') # rewrite removed DatasetCategory
        existing_categories_list = {item['name'] for item in existing_categories}

        if WDI_TAG_NAME not in existing_categories_list:
            parent_tag = Tag(name=WDI_TAG_NAME, is_bulk_import=True) # rewrite removed DatasetCategory
            parent_tag.save()
            logger.info("Inserting a category %s." % WDI_TAG_NAME.encode('utf8'))

        else:
            parent_tag = Tag.objects.get(name=WDI_TAG_NAME) # rewrite removed DatasetCategory

        existing_subcategories = Tag.objects.filter(parent_id=parent_tag).values('name') # rewrite removed DatasetSubcategory
        existing_subcategories_list = {item['name'] for item in existing_subcategories}

        wdi_categories_list = []

        for key, value in category_vars.items():
            wdi_categories_list.append(key)
            if key not in existing_subcategories_list:
                the_subcategory = Tag(name=key, parent_id=parent_tag) # rewrite removed DatasetSubcategory
                the_subcategory.save()
                logger.info("Inserting a subcategory %s." % key.encode('utf8'))

        cats_to_add = list(set(wdi_categories_list).difference(list(existing_subcategories_list)))

        existing_entities = Entity.objects.values('name')
        existing_entities_list = {item['name'] for item in existing_entities}

        country_tool_names = CountryName.objects.all()
        country_tool_names_dict = {}
        for each in country_tool_names:
            country_tool_names_dict[each.country_name.lower()] = each.owid_country

        country_name_entity_ref = {}  # this dict will hold the country names from excel and the appropriate entity object (this is used when saving the variables and their values)

        AdditionalCountryInfo.objects.filter(dataset=DATASET_NAMESPACE).delete()  # We will load new additional country data now

        row_number = 0
        for row in country_ws_rows:
            row_number += 1
            for cell in row:
                if row_number > 1:
                    column_number += 1
                    if column_number == 1:
                        country_code = cell.value
                    if column_number == 3:
                        country_name = cell.value
                    if column_number == 7:
                        country_special_notes = cell.value
                    if column_number == 8:
                        country_region = cell.value
                    if column_number == 9:
                        country_income_group = cell.value
                    if column_number == 24:
                        country_latest_census = cell.value
                    if column_number == 25:
                        country_latest_survey = cell.value
                    if column_number == 26:
                        country_recent_income_source = cell.value
                    if column_number == 30:
                        entity_info = AdditionalCountryInfo()
                        entity_info.country_code = country_code
                        entity_info.country_name = country_name
                        entity_info.country_wb_region = country_region
                        entity_info.country_wb_income_group = country_income_group
                        entity_info.country_special_notes = country_special_notes
                        entity_info.country_latest_census = country_latest_census
                        entity_info.country_latest_survey = country_latest_survey
                        entity_info.country_recent_income_source = country_recent_income_source
                        entity_info.save()
                        if country_tool_names_dict.get(unidecode.unidecode(country_name.lower()), 0):
                            newentity = Entity.objects.get(name=country_tool_names_dict[unidecode.unidecode(country_name.lower())].owid_name)
                        elif country_name in existing_entities_list:
                            newentity = Entity.objects.get(name=country_name)
                        else:
                            newentity = Entity(name=country_name, validated=False)
                            newentity.save()
                            logger.info("Inserting a country %s." % newentity.name.encode('utf8'))
                        country_name_entity_ref[country_code] = newentity

            column_number = 0

        insert_string = 'INSERT into data_values (value, year, entityId, variableId) VALUES (%s, %s, %s, %s)'  # this is used for constructing the query for mass inserting to the data_values table
        data_values_tuple_list = []

        total_values_tracker = 0
        dataset_id_oldname_list = []

        for category in wdi_categories_list:
            if category in cats_to_add:
                newdataset = Dataset(name='World Development Indicators - ' + category,
                                     description='This is a dataset imported by the automated fetcher',
                                     namespace=DATASET_NAMESPACE)
                newdataset.save()
                dataset_tag = DatasetTag(dataset_id=newdataset, tag_id=parent_tag)
                dataset_tag.save()
                dataset_id_oldname_list.append({'id': newdataset.pk, 'newname': newdataset.name, 'oldname': None})
                logger.info("Inserting a dataset %s." % newdataset.name.encode('utf8'))
            else:
                newdataset = Dataset.objects.get(name='World Development Indicators - ' + category)
                dataset_id_oldname_list.append({'id': newdataset.pk, 'newname': newdataset.name, 'oldname': newdataset.name})
            row_number = 0
            # TODO problem! This does multiple passes through the data
            # It will fail with the new early dereferenced generator
            for row in data_ws_rows:
                row_number += 1
                data_values = []
                for cell in row:
                    if row_number == 1:
                        if cell.value:
                            try:
                                last_available_year = int(cell.value)
                            except:
                                pass
                    if row_number > 1:
                        column_number += 1
                        if column_number == 1:
                            country_name = cell.value
                        if column_number == 2:
                            country_code = cell.value
                        if column_number == 3:
                            indicator_name = cell.value
                        if column_number == 4:
                            indicator_code = cell.value.upper().strip()
                        if column_number > 4 and column_number <= last_available_year - 1960 + 5:
                            if cell.value or cell.value == 0:
                                data_values.append({'value': cell.value, 'year': 1960 - 5 + column_number})
                        if column_number > 4 and column_number == last_available_year - 1960 + 5:
                            if len(data_values):
                                if indicator_code in category_vars[category]:
                                    total_values_tracker += len(data_values)
                                    if indicator_code in var_codes_to_add:
                                        source_description['additionalInfo'] = "Definitions and characteristics of countries and other territories: " + "https://ourworldindata.org" + reverse("servewdicountryinfo") + "\n"
                                        source_description['additionalInfo'] += "Limitations and exceptions:\n" + global_cat[indicator_code]['limitations'] + "\n" if global_cat[indicator_code]['limitations'] else ''
                                        source_description['additionalInfo'] += "Notes from original source:\n" + global_cat[indicator_code]['sourcenotes'] + "\n" if global_cat[indicator_code]['sourcenotes'] else ''
                                        source_description['additionalInfo'] += "General comments:\n" + global_cat[indicator_code]['comments'] + "\n" if global_cat[indicator_code]['comments'] else ''
                                        source_description['additionalInfo'] += "Statistical concept and methodology:\n" + global_cat[indicator_code]['concept'] if global_cat[indicator_code]['concept'] else ''
                                        source_description['additionalInfo'] += "Related source links:\n" + global_cat[indicator_code]['sourcelinks'] + "\n" if global_cat[indicator_code]['sourcelinks'] else ''
                                        source_description['additionalInfo'] += "Other web links:\n" + global_cat[indicator_code]['weblinks'] + "\n" if global_cat[indicator_code]['weblinks'] else ''
                                        source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                        if 'iea.org' in json.dumps(source_description).lower() or 'iea stat' in json.dumps(source_description).lower() or 'iea 2014' in json.dumps(source_description).lower():
                                            source_description['dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                        else:
                                            source_description['dataPublishedBy'] = 'World Bank – World Development Indicators'
                                        newsource = Source(name='World Bank – WDI: ' + global_cat[indicator_code]['name'],
                                                           description=json.dumps(source_description),
                                                           datasetId=newdataset)
                                        newsource.save()
                                        logger.info("Inserting a source %s." % newsource.name.encode('utf8'))
                                        global_cat[indicator_code]['source_object'] = newsource
                                        s_unit = extract_short_unit(global_cat[indicator_code]['unitofmeasure'])
                                        newvariable = Variable(name=global_cat[indicator_code]['name'],
                                                               unit=global_cat[indicator_code]['unitofmeasure'] if
                                                               global_cat[indicator_code]['unitofmeasure'] else '',
                                                               short_unit=s_unit,
                                                               description=global_cat[indicator_code]['description'],
                                                               code=indicator_code,
                                                               timespan='1960-' + str(last_available_year),
                                                               datasetId=newdataset,
                                                               sourceId=newsource)
                                        newvariable.save()
                                        global_cat[indicator_code]['variable_object'] = newvariable
                                        var_codes_to_add.remove(indicator_code)
                                        global_cat[indicator_code]['saved'] = True
                                        logger.info("Inserting a variable %s." % newvariable.name.encode('utf8'))
                                    else:
                                        if not global_cat[indicator_code]['saved']:
                                            newsource = Source.objects.get(name='World Bank – WDI: ' + Variable.objects.get(code=indicator_code, datasetId__in=Dataset.objects.filter(namespace=DATASET_NAMESPACE)).name)
                                            newsource.name = 'World Bank – WDI: ' + global_cat[indicator_code]['name']
                                            source_description['additionalInfo'] = "Definitions and characteristics of countries and other territories: " + "https://ourworldindata.org" + reverse("servewdicountryinfo") + "\n"
                                            source_description['additionalInfo'] += "Limitations and exceptions:\n" + global_cat[indicator_code]['limitations'] + "\n" if global_cat[indicator_code]['limitations'] else ''
                                            source_description['additionalInfo'] += "Notes from original source:\n" + global_cat[indicator_code]['sourcenotes'] + "\n" if global_cat[indicator_code]['sourcenotes'] else ''
                                            source_description['additionalInfo'] += "General comments:\n" + global_cat[indicator_code]['comments'] + "\n" if global_cat[indicator_code]['comments'] else ''
                                            source_description['additionalInfo'] += "Statistical concept and methodology:\n" + global_cat[indicator_code]['concept'] if global_cat[indicator_code]['concept'] else ''
                                            source_description['additionalInfo'] += "Related source links:\n" + global_cat[indicator_code]['sourcelinks'] + "\n" if global_cat[indicator_code]['sourcelinks'] else ''
                                            source_description['additionalInfo'] += "Other web links:\n" + global_cat[indicator_code]['weblinks'] + "\n" if global_cat[indicator_code]['weblinks'] else ''
                                            source_description['dataPublisherSource'] = global_cat[indicator_code]['source']
                                            if 'iea.org' in json.dumps(
                                                source_description).lower() or 'iea stat' in json.dumps(
                                                source_description).lower() or 'iea 2014' in json.dumps(
                                                source_description).lower():
                                                source_description[
                                                    'dataPublishedBy'] = 'International Energy Agency (IEA) via The World Bank'
                                            else:
                                                source_description[
                                                    'dataPublishedBy'] = 'World Bank – World Development Indicators'
                                            newsource.description=json.dumps(source_description)
                                            newsource.datasetId=newdataset
                                            newsource.save()
                                            logger.info("Updating the source %s." % newsource.name.encode('utf8'))
                                            s_unit = extract_short_unit(global_cat[indicator_code]['unitofmeasure'])
                                            newvariable = Variable.objects.get(code=indicator_code, datasetId__in=Dataset.objects.filter(namespace=DATASET_NAMESPACE))
                                            newvariable.name = global_cat[indicator_code]['name']
                                            newvariable.unit=global_cat[indicator_code]['unitofmeasure'] if global_cat[indicator_code]['unitofmeasure'] else ''
                                            newvariable.short_unit = s_unit
                                            newvariable.description=global_cat[indicator_code]['description']
                                            newvariable.timespan='1960-' + str(last_available_year)
                                            newvariable.datasetId=newdataset
                                            newvariable.sourceId=newsource
                                            newvariable.save()
                                            global_cat[indicator_code]['variable_object'] = newvariable
                                            logger.info("Updating the variable %s." % newvariable.name.encode('utf8'))
                                            global_cat[indicator_code]['saved'] = True
                                        else:
                                            newvariable = global_cat[indicator_code]['variable_object']
                                        if indicator_code not in newly_added_var_codes:
                                            if not deleted_indicators.get(indicator_code, 0):
                                                while DataValue.objects.filter(variableId__pk=newvariable.pk).first():
                                                    with connection.cursor() as c:
                                                        c.execute(
                                                                  'DELETE FROM %s WHERE variableId = %s LIMIT 10000;' %
                                                                  (DataValue._meta.db_table, newvariable.pk))
                                                deleted_indicators[indicator_code] = True
                                                logger.info("Deleting data values for the variable %s." % indicator_code.encode('utf8'))
                                    for i in range(0, len(data_values)):
                                        data_values_tuple_list.append((data_values[i]['value'], data_values[i]['year'],
                                                                       country_name_entity_ref[country_code].pk,
                                                                       newvariable.pk))
                                    if len(
                                        data_values_tuple_list) > 3000:  # insert when the length of the list goes over 3000
                                        with connection.cursor() as c:
                                            c.executemany(insert_string, data_values_tuple_list)
                                        logger.info("Dumping data values...")
                                        data_values_tuple_list = []
                column_number = 0
                if row_number % 10 == 0:
                    time.sleep(0.001)  # this is done in order to not keep the CPU busy all the time, the delay after each 10th row is 1 millisecond

        if len(data_values_tuple_list):  # insert any leftover data_values
            with connection.cursor() as c:
                c.executemany(insert_string, data_values_tuple_list)
            logger.info("Dumping data values...")

        # now deleting subcategories and datasets that are empty (that don't contain any variables), if any

        all_wdi_datasets = Dataset.objects.filter(namespace=DATASET_NAMESPACE)
        all_wdi_datasets_with_vars = Variable.objects.filter(datasetId__in=all_wdi_datasets).values(
            'datasetId').distinct()
        all_wdi_datasets_with_vars_dict = {item['datasetId'] for item in all_wdi_datasets_with_vars}

        for each in all_wdi_datasets:
            if each.pk not in all_wdi_datasets_with_vars_dict:
                cat_to_delete = each.subcategoryId
                logger.info("Deleting empty dataset %s." % each.name)
                logger.info("Deleting empty category %s." % cat_to_delete.name)
                each.delete()
                cat_to_delete.delete()

        newimport = ImportHistory(import_type=DATASET_NAMESPACE, import_time=timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                                  import_notes='Imported a total of %s data values.' % total_values_tracker,
                                  import_state=json.dumps(
                                      {'file_hash': file_checksum(WDI_DOWNLOADS_PATH + 'wdi.zip')}))
        newimport.save()

        # now exporting csvs to the repo
        # for dataset in dataset_id_oldname_list:
            # write_dataset_csv(dataset['id'], dataset['newname'], dataset['oldname'], 'wdi_fetcher', '')

print("--- %s seconds ---" % (time.time() - start_time))
